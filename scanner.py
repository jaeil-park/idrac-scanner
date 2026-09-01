"""
iDRAC 스캐너 + 일괄 설정 관리

스캔: Ping sweep → ARP/OUI → 포트 확인 → Redfish 핑거프린팅
설정: Redfish API / racadm CLI 프리셋 + 변수 치환 → 일괄 실행
"""

import collections
import concurrent.futures
import functools
import hashlib
import time
import ipaddress
import json
import os
import queue
import re
import shutil
import socket
import sqlite3
import ssl
import subprocess
import sys
import threading
import urllib.request
import uuid
from collections import Counter
from datetime import datetime

import requests as _req
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

from flask import (Flask, Response, jsonify, redirect, render_template,
                   request, session, stream_with_context, url_for)

try:                                   # 웹 SSH 콘솔용 (없어도 앱은 동작)
    from flask_sock import Sock
    _WS_AVAILABLE = True
except Exception:                      # pragma: no cover
    Sock = None
    _WS_AVAILABLE = False

IS_WINDOWS = sys.platform == "win32"

DB_PATH = os.environ.get("DB_PATH", "/data/known_devices.db")

CATEGORIES = {
    "idrac":   {"label": "iDRAC(등록)", "color": "#238636"},
    "infra":   {"label": "인프라",       "color": "#1c4186"},
    "test":    {"label": "테스트",       "color": "#9e6a03"},
    "exclude": {"label": "제외",         "color": "#484f58"},
}

# ── 내장 프리셋 정의 ──────────────────────────────────────────────

_BUILTIN_PRESETS = [
    {
        "name": "NTP 서버 설정",
        "description": "iDRAC NTP 서버 주소 및 시간 동기화 활성화",
        "category": "time",
        "engine": "both",
        "variables": json.dumps([
            {"name": "NTP1",  "label": "NTP 서버 1",         "type": "text",   "default": "pool.ntp.org",   "required": True},
            {"name": "NTP2",  "label": "NTP 서버 2 (선택)",   "type": "text",   "default": "time.google.com","required": False},
        ]),
        "commands": json.dumps([
            {"engine": "redfish", "method": "PATCH",
             "endpoint": "/redfish/v1/Managers/iDRAC.Embedded.1/Attributes",
             "body_template": '{"Attributes":{"NTPConfigGroup.1.NTPEnable":"Enabled","NTPConfigGroup.1.NTP1":"{{NTP1}}","NTPConfigGroup.1.NTP2":"{{NTP2}}"}}'},
            {"engine": "racadm", "commands": [
                "set iDRAC.NTPConfigGroup.NTPEnable 1",
                "set iDRAC.NTPConfigGroup.NTP1 {{NTP1}}",
                "set iDRAC.NTPConfigGroup.NTP2 {{NTP2}}",
            ]},
        ]),
    },
    {
        "name": "DNS 서버 설정",
        "description": "iDRAC IPv4 DNS 서버 주소 설정",
        "category": "network",
        "engine": "both",
        "variables": json.dumps([
            {"name": "DNS1", "label": "기본 DNS 서버", "type": "text", "default": "8.8.8.8",  "required": True},
            {"name": "DNS2", "label": "보조 DNS 서버", "type": "text", "default": "8.8.4.4",  "required": False},
        ]),
        "commands": json.dumps([
            {"engine": "redfish", "method": "PATCH",
             "endpoint": "/redfish/v1/Managers/iDRAC.Embedded.1/Attributes",
             "body_template": '{"Attributes":{"IPv4.1.DNSFromDHCP":"Disabled","IPv4.1.DNS1":"{{DNS1}}","IPv4.1.DNS2":"{{DNS2}}"}}'},
            {"engine": "racadm", "commands": [
                "set iDRAC.IPv4.DNSFromDHCP 0",
                "set iDRAC.IPv4.DNS1 {{DNS1}}",
                "set iDRAC.IPv4.DNS2 {{DNS2}}",
            ]},
        ]),
    },
    {
        "name": "iDRAC 사용자 비밀번호 변경",
        "description": "지정한 iDRAC 계정의 비밀번호 변경 (기본: root = ID 2)",
        "category": "user",
        "engine": "both",
        "variables": json.dumps([
            {"name": "USER_ID",      "label": "사용자 ID (숫자)",  "type": "text",     "default": "2",      "required": True},
            {"name": "NEW_PASSWORD", "label": "새 비밀번호",        "type": "password", "default": "",       "required": True},
        ]),
        "commands": json.dumps([
            {"engine": "redfish", "method": "PATCH",
             "endpoint": "/redfish/v1/AccountService/Accounts/{{USER_ID}}",
             "body_template": '{"Password":"{{NEW_PASSWORD}}"}'},
            {"engine": "racadm", "commands": [
                "set iDRAC.Users.{{USER_ID}}.Password {{NEW_PASSWORD}}",
            ]},
        ]),
    },
    {
        "name": "SNMP 커뮤니티 설정",
        "description": "SNMPv1/v2c 커뮤니티 문자열 및 트랩 대상 설정",
        "category": "alert",
        "engine": "both",
        "variables": json.dumps([
            {"name": "COMMUNITY",  "label": "커뮤니티 문자열", "type": "text", "default": "public",  "required": True},
            {"name": "TRAP_DEST",  "label": "트랩 대상 IP",    "type": "text", "default": "",        "required": False},
        ]),
        "commands": json.dumps([
            {"engine": "redfish", "method": "PATCH",
             "endpoint": "/redfish/v1/Managers/iDRAC.Embedded.1/Attributes",
             "body_template": '{"Attributes":{"SNMP.1.AgentEnable":"Enabled","SNMP.1.Community":"{{COMMUNITY}}","SNMP.1.TrapFormat":"SNMPv1","SNMP.1.AlertDestination1":"{{TRAP_DEST}}"}}'},
            {"engine": "racadm", "commands": [
                "set iDRAC.SNMP.AgentEnable 1",
                "set iDRAC.SNMP.Community {{COMMUNITY}}",
                "set iDRAC.SNMP.AlertDestination1 {{TRAP_DEST}}",
            ]},
        ]),
    },
    {
        "name": "Syslog 서버 설정",
        "description": "원격 Syslog 서버 주소 및 포트 설정",
        "category": "alert",
        "engine": "both",
        "variables": json.dumps([
            {"name": "SYSLOG_HOST", "label": "Syslog 서버 IP/호스트", "type": "text", "default": "", "required": True},
            {"name": "SYSLOG_PORT", "label": "포트",                   "type": "text", "default": "514","required": False},
        ]),
        "commands": json.dumps([
            {"engine": "redfish", "method": "PATCH",
             "endpoint": "/redfish/v1/Managers/iDRAC.Embedded.1/Attributes",
             "body_template": '{"Attributes":{"SysLog.1.SysLogEnable":"Enabled","SysLog.1.Server1":"{{SYSLOG_HOST}}","SysLog.1.Port":"{{SYSLOG_PORT}}"}}'},
            {"engine": "racadm", "commands": [
                "set iDRAC.SysLog.SysLogEnable 1",
                "set iDRAC.SysLog.Server1 {{SYSLOG_HOST}}",
                "set iDRAC.SysLog.Port {{SYSLOG_PORT}}",
            ]},
        ]),
    },
    {
        "name": "이메일 알림 설정",
        "description": "SMTP 서버를 통한 이메일 경보 설정",
        "category": "alert",
        "engine": "both",
        "variables": json.dumps([
            {"name": "SMTP_SERVER", "label": "SMTP 서버",      "type": "text", "default": "", "required": True},
            {"name": "FROM_EMAIL",  "label": "발신 이메일",    "type": "text", "default": "", "required": True},
            {"name": "TO_EMAIL",    "label": "수신 이메일",    "type": "text", "default": "", "required": True},
        ]),
        "commands": json.dumps([
            {"engine": "redfish", "method": "PATCH",
             "endpoint": "/redfish/v1/Managers/iDRAC.Embedded.1/Attributes",
             "body_template": '{"Attributes":{"RemoteHosts.1.SMTPServerIPAddress":"{{SMTP_SERVER}}","EmailAlert.1.Enable":"Enabled","EmailAlert.1.Address":"{{TO_EMAIL}}","EmailAlert.1.CustomMsg":"iDRAC Alert"}}'},
            {"engine": "racadm", "commands": [
                "set iDRAC.RemoteHosts.SMTPServerIPAddress {{SMTP_SERVER}}",
                "set iDRAC.EmailAlert.1.Enable 1",
                "set iDRAC.EmailAlert.1.Address {{TO_EMAIL}}",
            ]},
        ]),
    },
    {
        "name": "전원 정책 설정",
        "description": "서버 전원 관리 정책 설정",
        "category": "power",
        "engine": "both",
        "variables": json.dumps([
            {"name": "POWER_CAP",    "label": "전력 상한 (W, 0=비활성)", "type": "text",   "default": "0",    "required": False},
            {"name": "REDUNDANCY",   "label": "PSU 이중화 정책",          "type": "select",
             "options": ["Redundant", "NotRedundant", "PowerSavingMode"], "default": "Redundant", "required": False},
        ]),
        "commands": json.dumps([
            {"engine": "redfish", "method": "PATCH",
             "endpoint": "/redfish/v1/Chassis/System.Embedded.1/Power",
             "body_template": '{"PowerControl":[{"PowerLimit":{"LimitInWatts":{{POWER_CAP}},"LimitException":"HardPowerOff"}}]}'},
            {"engine": "racadm", "commands": [
                "set System.Power.RedundancyPolicy {{REDUNDANCY}}",
            ]},
        ]),
    },
    {
        "name": "BIOS 부팅 모드 (Redfish)",
        "description": "BootMode(UEFI/BIOS)·부팅 순서 재시도 설정 — racadm 없이 Redfish로 적용 (다음 재부팅 시)",
        "category": "bios",
        "engine": "redfish",
        "variables": json.dumps([
            {"name": "BOOT_MODE",      "label": "부팅 모드",        "type": "select",
             "options": ["Uefi", "Bios"], "default": "Uefi", "required": True},
            {"name": "BOOT_SEQ_RETRY", "label": "부팅 순서 재시도",  "type": "select",
             "options": ["Enabled", "Disabled"], "default": "Enabled", "required": False},
        ]),
        "commands": json.dumps([
            {"engine": "redfish", "method": "PATCH",
             "endpoint": "/redfish/v1/Systems/System.Embedded.1/Bios/Settings",
             "body_template": '{"Attributes":{"BootMode":"{{BOOT_MODE}}","BootSeqRetry":"{{BOOT_SEQ_RETRY}}"},"@Redfish.SettingsApplyTime":{"ApplyTime":"OnReset"}}'},
        ]),
    },
    {
        "name": "BIOS PXE 부팅 설정",
        "description": "PXE 네트워크 부팅 활성화 — Redfish(재부팅 시 적용) 또는 racadm",
        "category": "bios",
        "engine": "both",
        "variables": json.dumps([
            {"name": "PXE_NIC",  "label": "PXE NIC (예: NIC.Integrated.1-1-1)", "type": "text", "default": "NIC.Integrated.1-1-1", "required": True},
        ]),
        "commands": json.dumps([
            {"engine": "redfish", "method": "PATCH",
             "endpoint": "/redfish/v1/Systems/System.Embedded.1/Bios/Settings",
             "body_template": '{"Attributes":{"BootMode":"Bios","BootSeqRetry":"Enabled","PxeDev1EnDis":"Enabled","PxeDev1Interface":"{{PXE_NIC}}"},"@Redfish.SettingsApplyTime":{"ApplyTime":"OnReset"}}'},
            {"engine": "racadm", "commands": [
                "set BIOS.BiosBootSettings.BootMode Bios",
                "set BIOS.BiosBootSettings.BootSeqRetry Enabled",
                "jobqueue create BIOS.Setup.1-1",
            ]},
        ]),
    },
    {
        "name": "커스텀 racadm 명령",
        "description": "직접 racadm 명령어 입력 (줄바꿈으로 구분, 변수 {{VAR}} 사용 가능)",
        "category": "custom",
        "engine": "racadm",
        "variables": json.dumps([
            {"name": "CMD", "label": "racadm 명령어 (줄바꿈 구분)", "type": "textarea", "default": "get iDRAC.Info", "required": True},
        ]),
        "commands": json.dumps([
            {"engine": "racadm", "commands": ["{{CMD}}"]},
        ]),
    },
    {
        "name": "커스텀 Redfish PATCH",
        "description": "Redfish 엔드포인트와 JSON 본문 직접 지정",
        "category": "custom",
        "engine": "redfish",
        "variables": json.dumps([
            {"name": "ENDPOINT", "label": "Redfish 경로 (예: /redfish/v1/Managers/iDRAC.Embedded.1/Attributes)", "type": "text",     "default": "/redfish/v1/Managers/iDRAC.Embedded.1/Attributes", "required": True},
            {"name": "BODY",     "label": "JSON 본문",  "type": "textarea", "default": '{"Attributes":{}}', "required": True},
        ]),
        "commands": json.dumps([
            {"engine": "redfish", "method": "PATCH",
             "endpoint": "{{ENDPOINT}}",
             "body_template": "{{BODY}}"},
        ]),
    },
]


# ── DB 초기화 ─────────────────────────────────────────────────────

def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    with sqlite3.connect(DB_PATH) as c:
        c.executescript("""
            CREATE TABLE IF NOT EXISTS known_devices (
                ip          TEXT PRIMARY KEY,
                category    TEXT NOT NULL DEFAULT 'infra',
                label       TEXT NOT NULL DEFAULT '',
                note        TEXT NOT NULL DEFAULT '',
                mac         TEXT NOT NULL DEFAULT '',
                service_tag TEXT NOT NULL DEFAULT '',
                model       TEXT NOT NULL DEFAULT '',
                added_at    TEXT NOT NULL DEFAULT (datetime('now','localtime')),
                updated_at  TEXT NOT NULL DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS credentials (
                ip         TEXT PRIMARY KEY,
                username   TEXT NOT NULL DEFAULT 'root',
                password   TEXT NOT NULL DEFAULT '',
                note       TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS presets (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                name        TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                category    TEXT NOT NULL DEFAULT 'custom',
                engine      TEXT NOT NULL DEFAULT 'both',
                variables   TEXT NOT NULL DEFAULT '[]',
                commands    TEXT NOT NULL DEFAULT '[]',
                builtin     INTEGER NOT NULL DEFAULT 0,
                created_at  TEXT NOT NULL DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS exec_history (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                preset_name TEXT NOT NULL,
                engine      TEXT NOT NULL DEFAULT '',
                targets     TEXT NOT NULL DEFAULT '[]',
                variables   TEXT NOT NULL DEFAULT '{}',
                results     TEXT NOT NULL DEFAULT '{}',
                executed_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS scp_profiles (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                name        TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                content     TEXT NOT NULL DEFAULT '',
                component   TEXT NOT NULL DEFAULT 'ALL',
                created_at  TEXT NOT NULL DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS hw_info (
                ip         TEXT PRIMARY KEY,
                data       TEXT NOT NULL DEFAULT '{}',
                updated_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS app_meta (
                k TEXT PRIMARY KEY,
                v TEXT NOT NULL DEFAULT ''
            );
        """)
    # 기존 DB 컬럼 마이그레이션
    with sqlite3.connect(DB_PATH) as c:
        for col, defval in [("service_tag", "''"), ("model", "''")]:
            try:
                c.execute(f"ALTER TABLE known_devices ADD COLUMN {col} TEXT NOT NULL DEFAULT {defval}")
            except Exception:
                pass
    _seed_presets()


_BUILTIN_PRESETS_VERSION = 2   # 내장 프리셋 정의가 바뀌면 올릴 것 → 기동 시 재시드


def _seed_presets():
    with sqlite3.connect(DB_PATH) as c:
        row = c.execute("SELECT v FROM app_meta WHERE k='builtin_presets_version'").fetchone()
        ver = int(row[0]) if row and str(row[0]).isdigit() else 0
        have = c.execute("SELECT COUNT(*) FROM presets WHERE builtin=1").fetchone()[0]
        if have and ver >= _BUILTIN_PRESETS_VERSION:
            return
        # 내장 프리셋만 교체 (사용자 정의 builtin=0 프리셋은 보존)
        c.execute("DELETE FROM presets WHERE builtin=1")
        for p in _BUILTIN_PRESETS:
            c.execute("""
                INSERT INTO presets (name, description, category, engine, variables, commands, builtin)
                VALUES (?, ?, ?, ?, ?, ?, 1)
            """, (p["name"], p["description"], p["category"],
                  p["engine"], p["variables"], p["commands"]))
        c.execute("INSERT INTO app_meta (k, v) VALUES ('builtin_presets_version', ?) "
                  "ON CONFLICT(k) DO UPDATE SET v=excluded.v",
                  (str(_BUILTIN_PRESETS_VERSION),))


app = Flask(__name__)

# ── 인증 설정 ─────────────────────────────────────────────────────
_SECRET_KEY   = os.environ.get("SECRET_KEY", "")
app.secret_key = _SECRET_KEY.encode() if _SECRET_KEY else os.urandom(24)

AUTH_USER     = os.environ.get("AUTH_USER", "admin")
AUTH_PASSWORD = os.environ.get("AUTH_PASSWORD", "")   # 비어 있으면 인증 비활성화

# 웹 SSH 콘솔 (flask-sock)
sock = Sock(app) if _WS_AVAILABLE else None


def _auth_enabled() -> bool:
    return bool(AUTH_PASSWORD)


def require_auth(f):
    """개별 라우트용 데코레이터 (before_request 로 전체 적용)."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        return f(*args, **kwargs)
    return wrapper


@app.before_request
def _global_auth_check():
    """AUTH_PASSWORD가 설정된 경우 모든 요청에 인증을 요구."""
    if not _auth_enabled():
        return None
    if request.endpoint in ("login", "logout", "static"):
        return None
    if session.get("authenticated"):
        return None
    if request.path.startswith("/api/"):
        return jsonify({"error": "Unauthorized", "login": "/login"}), 401
    if request.path.startswith("/ws/"):
        return Response("Unauthorized", status=401)
    return redirect(url_for("login", next=request.full_path))


# ── 유틸 ──────────────────────────────────────────────────────────

DELL_OUIS = {
    "00:0B:DB", "00:14:22", "00:1A:A0", "00:1E:C9", "00:21:9B",
    "00:22:19", "00:23:AE", "00:24:E8", "00:26:B9",
    "14:18:77", "18:66:DA", "24:6E:96", "34:17:EB", "3C:2C:54",
    "44:A8:42", "54:BF:64", "78:2B:CB", "84:8F:69", "88:36:6C",
    "B0:83:FE", "B8:CA:3A", "D0:94:66", "EC:F4:BB", "F0:1F:AF",
    "F4:8E:38", "F8:BC:12",
}
IDRAC_PORTS = [443, 5900, 80, 623]
# 동일 MAC이 이 개수 이상 IP에서 나타나면 게이트웨이/라우터 MAC으로 판단
GATEWAY_MAC_THRESHOLD = 3

_SSL_CTX = ssl.create_default_context()
_SSL_CTX.check_hostname = False
_SSL_CTX.verify_mode = ssl.CERT_NONE

_IDRAC_PAGE_PATTERNS = [b"iDRAC", b"Integrated Dell Remote Access Controller"]


def normalize_mac(mac: str) -> str:
    mac = mac.upper().replace("-", ":").replace(".", ":")
    parts = mac.split(":")
    return ":".join(p.zfill(2) for p in parts) if len(parts) == 6 else mac


def is_dell_mac(mac: str) -> bool:
    norm = normalize_mac(mac)
    oui = ":".join(norm.split(":")[:3])
    return oui in DELL_OUIS


def sse(event_type: str, **kwargs) -> str:
    return f"data: {json.dumps({'type': event_type, **kwargs})}\n\n"


def substitute_vars(template: str, var_values: dict) -> str:
    """{{VAR_NAME}} 형태의 변수를 치환 (줄바꿈 명령은 분리해서 처리)"""
    result = template
    for k, v in var_values.items():
        result = result.replace(f"{{{{{k}}}}}", str(v))
    return result


def racadm_available() -> bool:
    return shutil.which("racadm") is not None


# ── 스캔 관련 ──────────────────────────────────────────────────────

def _ping_one(ip: str, timeout_ms: int = 400) -> bool:
    try:
        if IS_WINDOWS:
            cmd = ["ping", "-n", "1", "-w", str(timeout_ms), ip]
        else:
            cmd = ["ping", "-c", "1", "-W", "1", ip]
        result = subprocess.run(cmd, capture_output=True, timeout=timeout_ms / 1000 + 2)
        return result.returncode == 0
    except Exception:
        return False


def ping_sweep(network: str, max_workers: int = 300, timeout_ms: int = 400) -> list[str]:
    net = ipaddress.ip_network(network, strict=False)
    hosts = [str(ip) for ip in net.hosts()]
    alive = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(_ping_one, ip, timeout_ms): ip for ip in hosts}
        for fut in concurrent.futures.as_completed(futures):
            if fut.result():
                alive.append(futures[fut])
    return alive


def read_arp_cache() -> dict[str, str]:
    table: dict[str, str] = {}
    mac_re = re.compile(
        r"(\d{1,3}(?:\.\d{1,3}){3})\s+"
        r"([0-9a-fA-F]{2}(?:[-:][0-9a-fA-F]{2}){5})"
    )
    if IS_WINDOWS:
        try:
            out = subprocess.check_output(["arp", "-a"], text=True, timeout=5)
            for m in mac_re.finditer(out):
                ip, mac = m.group(1), normalize_mac(m.group(2))
                if not mac.startswith("FF:FF"):
                    table[ip] = mac
        except Exception:
            pass
    else:
        try:
            out = subprocess.check_output(["ip", "neigh"], text=True, timeout=5)
            for line in out.splitlines():
                parts = line.split()
                if len(parts) >= 5 and parts[3] == "lladdr":
                    ip, mac = parts[0], normalize_mac(parts[4])
                    if not mac.startswith("FF:FF"):
                        table[ip] = mac
        except Exception:
            pass
        if not table:
            try:
                out = subprocess.check_output(["arp", "-a"], text=True, timeout=5)
                for m in mac_re.finditer(out):
                    ip, mac = m.group(1), normalize_mac(m.group(2))
                    if not mac.startswith("FF:FF"):
                        table[ip] = mac
            except Exception:
                pass
    return table


def open_ports(ip: str, ports: list[int], timeout: float = 1.0) -> list[int]:
    result = []
    for port in ports:
        try:
            with socket.create_connection((ip, port), timeout=timeout):
                result.append(port)
        except Exception:
            pass
    return result


def _fetch_bytes(url: str, timeout: float = 5.0) -> bytes:
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 iDRACScanner"})
        with urllib.request.urlopen(req, timeout=timeout, context=_SSL_CTX) as r:
            return r.read(8192)
    except Exception:
        return b""


def _check_redfish(ip: str) -> tuple[bool, str]:
    import json as _json
    data = _fetch_bytes(f"https://{ip}/redfish/v1/", timeout=5.0)
    if not data:
        return False, ""
    try:
        obj = _json.loads(data)
        vendor  = str(obj.get("Vendor", "")).lower()
        product = str(obj.get("Product", "")).lower()
        name    = str(obj.get("Name", "")).lower()
        ver     = obj.get("RedfishVersion", "")
        if "dell" in vendor or "idrac" in product or "idrac" in name:
            ver_m = re.search(r"iDRAC\.Embedded\.(\d+)", str(obj))
            idrac_ver = f"iDRAC{ver_m.group(1)}" if ver_m else "iDRAC"
            return True, f"{idrac_ver} Redfish({ver})"
        if ver:
            return False, f"Redfish {ver} (non-Dell)"
    except Exception:
        pass
    return False, ""


def _load_fallback_creds() -> list:
    """IDRAC_FALLBACK_CREDS 환경변수(JSON)에서 자격증명 목록을 읽는다.
    미설정 시 기본값 [("root","calvin")] 만 반환."""
    raw = os.environ.get("IDRAC_FALLBACK_CREDS", "")
    if raw:
        try:
            parsed = json.loads(raw)
            return [tuple(c) for c in parsed if len(c) == 2]
        except Exception:
            pass
    return [("root", "calvin")]


def _fetch_system_info(ip: str) -> dict:
    """Redfish에서 서비스 태그(SKU)와 모델명을 가져온다.
    인증 없이 먼저 시도 → 저장된 자격증명 → 알려진 기본값 순으로 fallback."""
    result = {"service_tag": "", "model": ""}

    cred = _get_cred(ip)  # DB 저장값 or root/calvin
    _DEFAULT_CREDS = _load_fallback_creds()

    auth_list = [None, (cred["username"], cred["password"])]
    for dc in _DEFAULT_CREDS:
        if dc != (cred["username"], cred["password"]):
            auth_list.append(dc)

    paths = ["/redfish/v1/Systems/System.Embedded.1", "/redfish/v1/Systems/1"]

    for auth in auth_list:
        for path in paths:
            try:
                r = _req.get(f"https://{ip}{path}",
                             auth=auth, verify=False, timeout=6.0)
                if r.status_code == 200:
                    obj = r.json()
                    result["service_tag"] = str(obj.get("SKU") or "").strip()
                    result["model"]       = str(obj.get("Model") or "").strip()
                    if result["service_tag"] or result["model"]:
                        return result
            except Exception:
                pass
        # Systems 컬렉션 자동 탐색
        try:
            r = _req.get(f"https://{ip}/redfish/v1/Systems",
                         auth=auth, verify=False, timeout=6.0)
            if r.status_code == 200:
                for m in r.json().get("Members", []):
                    member_path = m.get("@odata.id", "")
                    if not member_path:
                        continue
                    r2 = _req.get(f"https://{ip}{member_path}",
                                  auth=auth, verify=False, timeout=6.0)
                    if r2.status_code == 200:
                        obj2 = r2.json()
                        result["service_tag"] = str(obj2.get("SKU") or "").strip()
                        result["model"]       = str(obj2.get("Model") or "").strip()
                        if result["service_tag"] or result["model"]:
                            return result
        except Exception:
            pass
    return result


# ── 하드웨어 상세 정보 (Redfish) ──────────────────────────────────

def _auth_candidates(ip: str) -> list:
    """해당 IP에 시도할 (user, pass) 목록 — 저장값 → 폴백 → 무인증."""
    cred = _get_cred(ip)
    cands: list = [(cred["username"], cred["password"])]
    for dc in _load_fallback_creds():
        t = tuple(dc)
        if t not in cands:
            cands.append(t)
    cands.append(None)
    return cands


def _resolve_auth(ip: str):
    """인증이 실제로 통하는 자격증명을 하나 찾아 반환 (없으면 None)."""
    test_paths = ["/redfish/v1/Managers/iDRAC.Embedded.1",
                  "/redfish/v1/Systems/System.Embedded.1"]
    for auth in _auth_candidates(ip):
        for tp in test_paths:
            try:
                r = _req.get(f"https://{ip}{tp}", auth=auth, verify=False, timeout=6)
                if r.status_code == 200:
                    return auth
                if r.status_code in (401, 403):
                    break
            except Exception:
                pass
    return None


def _fetch_hw_detail(ip: str, want_fw: bool = False) -> dict:
    """Redfish로 CPU/메모리/디스크/NIC/PSU/펌웨어 정보를 수집."""
    out: dict = {
        "ip": ip, "ok": False, "error": "",
        "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "system": {}, "processors": [], "memory": [], "storage": [],
        "network": [], "power": {}, "idrac": {}, "firmware": [],
    }
    auth = _resolve_auth(ip)
    if auth is None:
        out["error"] = "인증 실패 — 일괄 설정 페이지에서 IP별 자격증명을 저장하세요."
        return out

    def g(path: str):
        if not path:
            return None
        try:
            r = _req.get(f"https://{ip}{path}", auth=auth, verify=False, timeout=12)
            return r.json() if r.status_code == 200 else None
        except Exception:
            return None

    def health(obj):
        s = (obj or {}).get("Status", {}) or {}
        return s.get("Health", "") or s.get("HealthRollup", "") or ""

    # ── System ──
    sysobj = g("/redfish/v1/Systems/System.Embedded.1") or g("/redfish/v1/Systems/1")
    if not sysobj:
        for m in (g("/redfish/v1/Systems") or {}).get("Members", []):
            sysobj = g(m.get("@odata.id", ""))
            if sysobj:
                break
    sys_path = (sysobj or {}).get("@odata.id", "/redfish/v1/Systems/System.Embedded.1")
    if sysobj:
        ps = sysobj.get("ProcessorSummary", {}) or {}
        ms = sysobj.get("MemorySummary", {}) or {}
        out["system"] = {
            "model":        sysobj.get("Model", ""),
            "manufacturer": sysobj.get("Manufacturer", ""),
            "service_tag":  sysobj.get("SKU", "") or sysobj.get("SerialNumber", ""),
            "hostname":     sysobj.get("HostName", ""),
            "bios_version": sysobj.get("BiosVersion", ""),
            "power_state":  sysobj.get("PowerState", ""),
            "health":       health(sysobj),
            "indicator_led": sysobj.get("IndicatorLED", ""),
            "cpu_count":    ps.get("Count", ""),
            "cpu_model":    (ps.get("Model", "") or "").strip(),
            "mem_total_gib": ms.get("TotalSystemMemoryGiB", ""),
        }

    # ── Processors ──
    for m in (g(sys_path + "/Processors") or {}).get("Members", [])[:16]:
        p = g(m.get("@odata.id", ""))
        if not p:
            continue
        if p.get("ProcessorType", "CPU") not in ("CPU", "", None):
            continue
        out["processors"].append({
            "id":        p.get("Id", ""),
            "model":     (p.get("Model", "") or "").strip(),
            "cores":     p.get("TotalCores", ""),
            "threads":   p.get("TotalThreads", ""),
            "speed_mhz": p.get("MaxSpeedMHz", ""),
            "health":    health(p),
            "state":     (p.get("Status", {}) or {}).get("State", ""),
        })

    # ── Memory ──
    for m in (g(sys_path + "/Memory") or {}).get("Members", [])[:64]:
        d = g(m.get("@odata.id", ""))
        if not d:
            continue
        st = d.get("Status", {}) or {}
        cap = d.get("CapacityMiB")
        if not cap and st.get("State") == "Absent":
            continue
        speed = d.get("OperatingSpeedMhz", "")
        if not speed:
            allowed = d.get("AllowedSpeedsMHz") or []
            speed = allowed[0] if allowed else ""
        out["memory"].append({
            "locator":      d.get("DeviceLocator", "") or d.get("Id", ""),
            "capacity_gib": round(cap / 1024, 1) if cap else "",
            "speed_mhz":    speed,
            "type":         d.get("MemoryDeviceType", ""),
            "manufacturer": (d.get("Manufacturer", "") or "").strip(),
            "part_number":  (d.get("PartNumber", "") or "").strip(),
            "serial":       (d.get("SerialNumber", "") or "").strip(),
            "health":       st.get("Health", ""),
            "state":        st.get("State", ""),
        })

    # ── Storage ──
    for m in (g(sys_path + "/Storage") or {}).get("Members", [])[:16]:
        ctrl = g(m.get("@odata.id", ""))
        if not ctrl:
            continue
        scs = ctrl.get("StorageControllers", []) or [{}]
        cinfo = {
            "name":     ctrl.get("Name", ""),
            "model":    scs[0].get("Model", "") or ctrl.get("Name", ""),
            "firmware": scs[0].get("FirmwareVersion", ""),
            "health":   health(ctrl),
            "drives":   [],
        }
        for dref in ctrl.get("Drives", [])[:64]:
            dr = g(dref.get("@odata.id", ""))
            if not dr:
                continue
            cap = dr.get("CapacityBytes") or 0
            cinfo["drives"].append({
                "name":        dr.get("Name", ""),
                "capacity_gb": round(cap / 1e9, 1) if cap else "",
                "media":       dr.get("MediaType", ""),
                "proto":       dr.get("Protocol", ""),
                "model":       (dr.get("Model", "") or "").strip(),
                "serial":      (dr.get("SerialNumber", "") or "").strip(),
                "health":      health(dr),
                "state":       (dr.get("Status", {}) or {}).get("State", ""),
            })
        out["storage"].append(cinfo)

    # ── Network (EthernetInterfaces) ──
    for m in (g(sys_path + "/EthernetInterfaces") or {}).get("Members", [])[:32]:
        n = g(m.get("@odata.id", ""))
        if not n:
            continue
        out["network"].append({
            "id":         n.get("Id", ""),
            "name":       n.get("Name", ""),
            "mac":        n.get("MACAddress", "") or n.get("PermanentMACAddress", ""),
            "speed_mbps": n.get("SpeedMbps", ""),
            "link":       n.get("LinkStatus", ""),
            "health":     health(n),
        })

    # ── Power / PSU ──
    pw = (g("/redfish/v1/Chassis/System.Embedded.1/Power")
          or g("/redfish/v1/Chassis/1/Power") or {})
    psus = []
    for p in pw.get("PowerSupplies", []):
        psus.append({
            "name":       p.get("Name", ""),
            "model":      p.get("Model", ""),
            "serial":     p.get("SerialNumber", ""),
            "firmware":   p.get("FirmwareVersion", ""),
            "capacity_w": p.get("PowerCapacityWatts", ""),
            "input_w":    p.get("PowerInputWatts", ""),
            "health":     health(p),
            "state":      (p.get("Status", {}) or {}).get("State", ""),
        })
    pcs = pw.get("PowerControl") or [{}]
    out["power"] = {
        "supplies":   psus,
        "consumed_w": pcs[0].get("PowerConsumedWatts", "") if pcs else "",
        "capacity_w": pcs[0].get("PowerCapacityWatts", "") if pcs else "",
    }

    # ── iDRAC Manager ──
    mgr = (g("/redfish/v1/Managers/iDRAC.Embedded.1")
           or g("/redfish/v1/Managers/1") or {})
    out["idrac"] = {
        "firmware_version": mgr.get("FirmwareVersion", ""),
        "model":            mgr.get("Model", ""),
        "datetime":         mgr.get("DateTime", ""),
        "health":           health(mgr),
    }

    # ── Firmware Inventory (선택 — 느림) ──
    if want_fw:
        seen = set()
        for m in (g("/redfish/v1/UpdateService/FirmwareInventory") or {}).get("Members", [])[:250]:
            fid = m.get("@odata.id", "")
            if "Installed" not in fid:
                continue
            f = g(fid)
            if not f or not f.get("Version"):
                continue
            key = (f.get("Name", ""), f.get("Version", ""))
            if key in seen:
                continue
            seen.add(key)
            out["firmware"].append({
                "name":       f.get("Name", ""),
                "version":    f.get("Version", ""),
                "updateable": f.get("Updateable", ""),
            })
        out["firmware"] = sorted(out["firmware"], key=lambda x: x["name"])[:120]

    out["ok"] = True
    return out


# ── iDRAC 펌웨어 업데이트 (Redfish) ──────────────────────────────

_FW_TMP_DIR = os.environ.get("FW_TMP_DIR", "/tmp/idrac_fw")
# NAS 등 서버에서 접근 가능한 펌웨어 보관 디렉토리 (컨테이너에 마운트)
_FW_LIB_DIR = os.environ.get("FW_LIB_DIR", "")
_FW_EXTS = (".exe", ".d9", ".d7", ".bin", ".pm", ".usc", ".ph", ".rpm")


def _fw_lib_root() -> str:
    return os.path.realpath(_FW_LIB_DIR) if _FW_LIB_DIR else ""


def _resolve_fw_lib_path(rel: str) -> str:
    """라이브러리 디렉토리 내부의 파일만 허용 (경로 탈출 차단)."""
    root = _fw_lib_root()
    if not root or not rel:
        return ""
    target = os.path.realpath(os.path.join(root, rel.lstrip("/\\")))
    if target != root and not target.startswith(root + os.sep):
        return ""
    return target if os.path.isfile(target) else ""


def _run_fw_update(q: queue.Queue, payload: dict, file_path: str,
                   file_name: str, cleanup: bool = True):
    """백그라운드 — 대상별로 Redfish 펌웨어 업로드/설치 후 Task 폴링."""
    targets     = payload.get("targets", [])
    use_common  = payload.get("use_common_cred", False)
    common_cred = payload.get("common_cred", {})
    apply_time  = payload.get("apply_time", "Immediate")     # Immediate | OnReset
    image_uri   = (payload.get("image_uri") or "").strip()
    xfer_proto  = payload.get("transfer_protocol", "HTTP")
    share_user  = payload.get("share_user", "")
    share_pass  = payload.get("share_pass", "")

    for ip in targets:
        cred = common_cred if use_common else _get_cred(ip)
        auth = (cred["username"], cred["password"])
        q.put(sse("target_start", ip=ip))
        ok_final = False
        try:
            us = {}
            try:
                r = _req.get(f"https://{ip}/redfish/v1/UpdateService",
                             auth=auth, verify=False, timeout=15)
                if r.status_code == 200:
                    us = r.json()
            except Exception:
                pass

            if image_uri:
                act = (us.get("Actions", {}) or {}).get("#UpdateService.SimpleUpdate", {}) or {}
                target = act.get("target",
                    "/redfish/v1/UpdateService/Actions/UpdateService.SimpleUpdate")
                body = {"ImageURI": image_uri, "TransferProtocol": xfer_proto,
                        "@Redfish.OperationApplyTime": apply_time}
                if share_user:
                    body["Username"] = share_user
                if share_pass:
                    body["Password"] = share_pass
                resp = _req.post(f"https://{ip}{target}", auth=auth, json=body,
                                 verify=False, timeout=90)
                step = {"engine": "fw", "cmd": f"SimpleUpdate ← {image_uri}",
                        "ok": resp.status_code in (200, 202),
                        "output": f"HTTP {resp.status_code} {resp.text[:200]}"}
            else:
                push_uri = us.get("MultipartHTTPPushUri") or "/redfish/v1/UpdateService/MultipartUpload"
                params = {"Targets": [], "@Redfish.OperationApplyTime": apply_time}
                with open(file_path, "rb") as fh:
                    files = {
                        "UpdateParameters": ("params.json", json.dumps(params), "application/json"),
                        "UpdateFile": (file_name, fh, "application/octet-stream"),
                    }
                    resp = _req.post(f"https://{ip}{push_uri}", auth=auth, files=files,
                                     verify=False, timeout=900)
                step = {"engine": "fw", "cmd": f"업로드 {file_name} (apply={apply_time})",
                        "ok": resp.status_code in (200, 202),
                        "output": f"HTTP {resp.status_code} {resp.text[:200]}"}

            q.put(sse("step_done", ip=ip, step=step))
            ok_final = step["ok"]

            loc = resp.headers.get("Location", "")
            job_id = loc.rstrip("/").split("/")[-1] if loc else ""
            if job_id and step["ok"]:
                for _ in range(180):            # 최대 ~30분
                    time.sleep(10)
                    try:
                        jr = _req.get(f"https://{ip}/redfish/v1/TaskService/Tasks/{job_id}",
                                      auth=auth, verify=False, timeout=20)
                        if jr.status_code == 404:
                            jr = _req.get(f"https://{ip}/redfish/v1/Managers/iDRAC.Embedded.1/Oem/Dell/Jobs/{job_id}",
                                          auth=auth, verify=False, timeout=20)
                        if jr.status_code != 200:
                            continue
                        t = jr.json()
                        state = t.get("TaskState") or t.get("JobState") or "Running"
                        pct = t.get("PercentComplete", 0)
                        if state in ("Completed", "Exception", "Killed", "Cancelled", "Failed"):
                            ok_final = state == "Completed"
                            msgs = " | ".join(m.get("Message", "")
                                              for m in t.get("Messages", []))
                            q.put(sse("step_done", ip=ip, step={
                                "engine": "fw", "cmd": f"작업 {job_id}",
                                "ok": ok_final, "output": f"{state} — {msgs}"}))
                            break
                        q.put(sse("step_done", ip=ip, step={
                            "engine": "fw", "cmd": f"진행 중 ({pct}%)",
                            "ok": True, "output": state}))
                    except Exception:
                        continue
        except Exception as e:
            q.put(sse("step_done", ip=ip, step={
                "engine": "fw", "cmd": "오류", "ok": False, "output": str(e)}))
        q.put(sse("target_done", ip=ip, ok=ok_final))

    q.put(sse("done", ts=datetime.now().strftime("%H:%M:%S")))
    q.put(None)
    if cleanup:
        try:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
        except Exception:
            pass


def fingerprint_idrac(ip: str) -> tuple[bool, str, str, str]:
    """(is_idrac, note, service_tag, model) 반환"""
    ok, note = _check_redfish(ip)
    if ok:
        info = _fetch_system_info(ip)
        return True, note, info["service_tag"], info["model"]
    if "Redfish" in note and "non-Dell" in note:
        return False, note, "", ""
    for scheme in ("https", "http"):
        data = _fetch_bytes(f"{scheme}://{ip}/")
        if not data:
            continue
        for pat in _IDRAC_PAGE_PATTERNS:
            if pat in data:
                ver_m = re.search(rb"iDRAC\s*(\d+)", data, re.I)
                ver = ver_m.group(1).decode() if ver_m else ""
                info = _fetch_system_info(ip)
                return True, f"iDRAC{ver} ({scheme}:page)", info["service_tag"], info["model"]
    return False, "응답 있음 (iDRAC 아님)", "", ""


def scan_generator(network: str, mode: str):
    yield sse("start", network=network, mode=mode, ts=datetime.now().strftime("%H:%M:%S"))
    try:
        net = ipaddress.ip_network(network, strict=False)
    except ValueError as e:
        yield sse("error", msg=f"잘못된 서브넷: {e}")
        return

    total_hosts = net.num_addresses - 2
    found: list[dict] = []

    if mode in ("arp", "both"):
        yield sse("progress", step="ping", percent=5, msg=f"Ping 스윕 중… ({total_hosts}개 호스트)")
        alive = ping_sweep(network)
        yield sse("progress", step="arp", percent=40, msg=f"ARP 캐시 읽는 중… (응답 {len(alive)}개)")
        arp = read_arp_cache()

        # 동일 MAC을 여러 IP가 공유하면 게이트웨이/라우터 MAC — OUI 필터 신뢰 불가
        mac_counts = Counter(arp.get(ip, "") for ip in alive if arp.get(ip, ""))
        gateway_macs = {mac for mac, cnt in mac_counts.items()
                        if cnt >= GATEWAY_MAC_THRESHOLD and mac}

        if gateway_macs:
            yield sse("warning", msg=(
                f"⚠ 게이트웨이/라우터 MAC {len(gateway_macs)}개 감지 "
                f"(동일 MAC이 {GATEWAY_MAC_THRESHOLD}개 이상 IP에서 출현) — "
                "해당 IP들은 포트 스캔+Redfish 방식으로 재탐지합니다."
            ))

        # 게이트웨이 MAC이 아닌 IP만 Dell OUI로 필터 (L2 직결 장비)
        dell_ips_direct = [
            ip for ip in alive
            if is_dell_mac(arp.get(ip, "")) and arp.get(ip, "") not in gateway_macs
        ]
        # 게이트웨이 뒤에 있어 실제 MAC 미확인인 살아있는 IP → 포트 스캔으로 탐지
        gateway_alive  = [ip for ip in alive if arp.get(ip, "") in gateway_macs]
        others         = [ip for ip in alive if ip not in dell_ips_direct and ip not in gateway_alive]

        yield sse("progress", step="filter", percent=50,
                   msg=f"Dell 직결 {len(dell_ips_direct)}개 · 라우터 경유 {len(gateway_alive)}개 확인 중…")

        def check_dell(ip):
            mac   = arp.get(ip, "")
            ports = open_ports(ip, IDRAC_PORTS)
            is_id, note, svc_tag, model = fingerprint_idrac(ip) if 443 in ports or 80 in ports else (False, "", "", "")
            return {"ip": ip, "mac": mac, "open_ports": ports,
                    "is_idrac": is_id, "note": note, "method": "ARP+OUI", "is_dell": True,
                    "service_tag": svc_tag, "model": model}


        def _save_svc(res):
            if res.get("is_idrac") and (res.get("service_tag") or res.get("model")):
                try:
                    with sqlite3.connect(DB_PATH) as _c:
                        _c.execute(
                            "UPDATE known_devices SET service_tag=?, model=?, updated_at=datetime('now','localtime')"
                            " WHERE ip=? AND (service_tag='' OR model='')",
                            (res.get("service_tag",""), res.get("model",""), res["ip"])
                        )
                except Exception:
                    pass

        # L2 직결 Dell 장비 처리
        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
            for i, res in enumerate(ex.map(check_dell, dell_ips_direct), 1):
                found.append(res)
                _save_svc(res)
                yield sse("found", host=res, percent=50 + int(20 * i / max(len(dell_ips_direct), 1)))

        # 게이트웨이 경유 호스트 처리 (포트 스캔)
        if gateway_alive:
            # 1단계: 포트 확인만 병렬로 (세션 불필요, 빠름)
            yield sse("progress", step="gateway", percent=70,
                       msg=f"게이트웨이 경유 {len(gateway_alive)}개 포트 스캔 중…")

            def port_only(ip):
                ports = open_ports(ip, IDRAC_PORTS + [623], timeout=1.5)
                return (ip, ports)

            with concurrent.futures.ThreadPoolExecutor(max_workers=100) as ex:
                port_results = list(ex.map(port_only, gateway_alive))

            # 2단계: 포트 열린 호스트만 iDRAC 핑거프린팅 (세션 경합 방지 위해 소수 동시)
            candidates = [(ip, ports) for ip, ports in port_results if ports]
            yield sse("progress", step="gateway_fp", percent=80,
                       msg=f"포트 열린 {len(candidates)}개 iDRAC 확인 중…")

            def fingerprint_gw(item):
                ip, ports = item
                is_id, note, svc_tag, model = (
                    fingerprint_idrac(ip) if (443 in ports or 80 in ports)
                    else (False, "", "", "")
                )
                real_mac = read_arp_cache().get(ip, "")
                return {"ip": ip, "mac": real_mac, "open_ports": ports,
                        "is_idrac": is_id, "note": note, "method": "PortScan(GW)",
                        "is_dell": is_id,
                        "service_tag": svc_tag, "model": model}

            with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
                results_gw = list(ex.map(fingerprint_gw, candidates))

            for i, res in enumerate(results_gw, 1):
                if res["is_idrac"] or res["open_ports"]:
                    found.append(res)
                    _save_svc(res)
                    yield sse("found", host=res,
                               percent=80 + int(15 * i / max(len(candidates), 1)))

        for ip in others:
            yield sse("other", host={"ip": ip, "mac": arp.get(ip, ""), "is_dell": False})

    if mode in ("port", "both"):
        already = {h["ip"] for h in found}
        hosts   = [str(ip) for ip in net.hosts() if str(ip) not in already]
        yield sse("progress", step="portscan", percent=60 if mode == "both" else 10,
                   msg=f"포트 스캔 중… ({len(hosts)}개, 443/623/5900)")

        lock = threading.Lock()

        def port_scan_host(ip):
            ports = open_ports(ip, [443, 5900, 623, 80], timeout=1.5)
            if not ports:
                return None
            is_id, note, svc_tag, model = fingerprint_idrac(ip) if 443 in ports or 80 in ports else (False, "", "", "")
            mac = read_arp_cache().get(ip, "")
            return {"ip": ip, "mac": mac, "open_ports": ports,
                    "is_idrac": is_id, "note": note, "method": "PortScan",
                    "is_dell": is_dell_mac(mac) if mac else False,
                    "service_tag": svc_tag, "model": model}

        def scan_batch(host_list):
            with concurrent.futures.ThreadPoolExecutor(max_workers=300) as ex:
                for res in ex.map(port_scan_host, host_list):
                    if res and (res["is_idrac"] or res["is_dell"] or 623 in res["open_ports"]):
                        with lock:
                            found.append(res)
                        yield res

        for res in scan_batch(hosts):
            yield sse("found", host=res, percent=90)

    # service_tag 미획득 iDRAC 장비를 순차적으로 재시도 (세션 완전 해소 후)
    retry_targets = [h for h in found if h.get("is_idrac") and not h.get("service_tag")]
    if retry_targets:
        time.sleep(2)   # 병렬 세션이 모두 닫힐 때까지 대기
        yield sse("progress", step="retry", percent=97,
                   msg=f"서비스 태그 재조회 중… ({len(retry_targets)}개)")
        for h in retry_targets:
            try:
                info = _fetch_system_info(h["ip"])
                if info.get("service_tag") or info.get("model"):
                    h["service_tag"] = info["service_tag"]
                    h["model"]       = info["model"]
                    try:
                        with sqlite3.connect(DB_PATH) as _c:
                            _c.execute(
                                "UPDATE known_devices SET service_tag=?, model=?, updated_at=datetime('now','localtime')"
                                " WHERE ip=? AND (service_tag='' OR model='')",
                                (info["service_tag"], info["model"], h["ip"])
                            )
                    except Exception:
                        pass
                    yield sse("service_tag_update", ip=h["ip"],
                               service_tag=h["service_tag"], model=h["model"])
            except Exception:
                pass

    # 같은 MAC을 공유하는 IP 그룹화 — 단, 게이트웨이 MAC은 제외
    found_mac_counts = Counter(h.get("mac","") for h in found if h.get("mac",""))
    found_gateway_macs = {mac for mac, cnt in found_mac_counts.items() if cnt >= GATEWAY_MAC_THRESHOLD and mac}
    mac_to_idrac: dict[str, str] = {}
    for h in found:
        if h.get("is_idrac") and h.get("mac") and h["mac"] not in found_gateway_macs:
            mac_to_idrac[h["mac"]] = h["ip"]
    shared_updates = []
    for h in found:
        mac = h.get("mac", "")
        if mac and not h.get("is_idrac") and mac in mac_to_idrac:
            h["parent_idrac"] = mac_to_idrac[mac]
            shared_updates.append({"ip": h["ip"], "parent_idrac": h["parent_idrac"]})
    if shared_updates:
        yield sse("shared_host_update", updates=shared_updates)

    yield sse("done", total_found=len(found),
               idrac_count=sum(1 for h in found if h.get("is_idrac")),
               dell_count=sum(1 for h in found if h.get("is_dell")),
               ts=datetime.now().strftime("%H:%M:%S"))


# ── 알려진 장비 헬퍼 ──────────────────────────────────────────────

def _known_map() -> dict[str, dict]:
    try:
        with sqlite3.connect(DB_PATH) as c:
            rows = c.execute(
                "SELECT ip, category, label, note, mac, added_at,"
                " COALESCE(service_tag,''), COALESCE(model,'') FROM known_devices"
            ).fetchall()
        return {r[0]: {"ip": r[0], "category": r[1], "label": r[2],
                        "note": r[3], "mac": r[4], "added_at": r[5],
                        "service_tag": r[6], "model": r[7]} for r in rows}
    except Exception:
        return {}


def _get_cred(ip: str) -> dict:
    """IP별 자격증명 조회, 없으면 기본값(root/calvin)"""
    try:
        with sqlite3.connect(DB_PATH) as c:
            row = c.execute(
                "SELECT username, password FROM credentials WHERE ip=?", (ip,)
            ).fetchone()
        if row:
            return {"username": row[0], "password": row[1]}
    except Exception:
        pass
    return {"username": "root", "password": "calvin"}


# ── 일괄 실행 엔진 ────────────────────────────────────────────────

def _exec_redfish_cmd(ip: str, cred: dict, cmd: dict, var_values: dict) -> dict:
    endpoint = substitute_vars(cmd.get("endpoint", ""), var_values)
    method   = cmd.get("method", "GET").upper()
    body_str = substitute_vars(cmd.get("body_template", "{}"), var_values)
    url = f"https://{ip}{endpoint}"
    try:
        body = json.loads(body_str)
    except Exception:
        return {"ok": False, "error": f"JSON 파싱 오류: {body_str[:80]}"}
    try:
        r = _req.request(
            method, url,
            auth=(cred["username"], cred["password"]),
            json=body, verify=False, timeout=20,
        )
        ok = r.status_code in (200, 201, 202, 204)
        return {"ok": ok, "status": r.status_code, "response": r.text[:300]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _exec_racadm_cmd(ip: str, cred: dict, cmd: dict, var_values: dict) -> list[dict]:
    results = []
    raw_cmds = cmd.get("commands", [])
    for line in raw_cmds:
        expanded = substitute_vars(line, var_values)
        # 변수가 줄바꿈을 포함할 수 있으므로 분리
        for sub_line in expanded.splitlines():
            sub_line = sub_line.strip()
            if not sub_line:
                continue
            full_cmd = ["racadm", "-r", ip, "-u", cred["username"],
                        "-p", cred["password"], "--nocertwarn"] + sub_line.split()
            try:
                proc = subprocess.run(full_cmd, capture_output=True, text=True, timeout=30)
                results.append({
                    "cmd": sub_line,
                    "ok": proc.returncode == 0,
                    "output": (proc.stdout + proc.stderr).strip()[:300],
                })
            except subprocess.TimeoutExpired:
                results.append({"cmd": sub_line, "ok": False, "output": "타임아웃 (30s)"})
            except FileNotFoundError:
                results.append({"cmd": sub_line, "ok": False, "output": "racadm 바이너리 없음"})
            except Exception as e:
                results.append({"cmd": sub_line, "ok": False, "output": str(e)})
    return results


_active_jobs: dict[str, queue.Queue] = {}


def _run_execution(q: queue.Queue, payload: dict):
    """백그라운드 스레드 — 실행 결과를 SSE 메시지 문자열로 queue에 적재"""
    preset_id   = payload.get("preset_id")
    targets     = payload.get("targets", [])
    var_values  = payload.get("variables", {})
    engine_sel  = payload.get("engine", "both")   # redfish / racadm / both
    use_common  = payload.get("use_common_cred", False)
    common_cred = payload.get("common_cred", {})

    try:
        with sqlite3.connect(DB_PATH) as c:
            row = c.execute("SELECT name, engine, commands FROM presets WHERE id=?",
                            (preset_id,)).fetchone()
        if not row:
            q.put(sse("error", msg="프리셋을 찾을 수 없습니다"))
            return
        preset_name, preset_engine, commands_json = row
        commands = json.loads(commands_json)
    except Exception as e:
        q.put(sse("error", msg=str(e)))
        q.put(None)
        return

    all_results = {}

    for ip in targets:
        cred = common_cred if use_common else _get_cred(ip)
        ip_results = []
        q.put(sse("target_start", ip=ip))

        for cmd in commands:
            cmd_engine = cmd.get("engine", "racadm")
            if engine_sel != "both" and cmd_engine != engine_sel:
                continue

            if cmd_engine == "redfish":
                res = _exec_redfish_cmd(ip, cred, cmd, var_values)
                step = {"engine": "redfish", "endpoint": cmd.get("endpoint", ""), **res}
                ip_results.append(step)
                q.put(sse("step_done", ip=ip, step=step))

            elif cmd_engine == "racadm":
                for r in _exec_racadm_cmd(ip, cred, cmd, var_values):
                    step = {"engine": "racadm", **r}
                    ip_results.append(step)
                    q.put(sse("step_done", ip=ip, step=step))

        overall_ok = all(s.get("ok", False) for s in ip_results) if ip_results else False
        all_results[ip] = {"ok": overall_ok, "steps": ip_results}
        q.put(sse("target_done", ip=ip, ok=overall_ok))

    # 이력 저장
    try:
        with sqlite3.connect(DB_PATH) as c:
            c.execute("""
                INSERT INTO exec_history (preset_name, engine, targets, variables, results)
                VALUES (?, ?, ?, ?, ?)
            """, (preset_name, engine_sel, json.dumps(targets),
                  json.dumps(var_values), json.dumps(all_results)))
    except Exception:
        pass

    q.put(sse("done", results=all_results, ts=datetime.now().strftime("%H:%M:%S")))
    q.put(None)


# ── Flask 라우트 ──────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if not _auth_enabled():
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        user = request.form.get("username", "").strip()
        pw   = request.form.get("password", "")
        if user == AUTH_USER and pw == AUTH_PASSWORD:
            session["authenticated"] = True
            nxt = request.args.get("next") or url_for("index")
            # 오픈 리다이렉트 방지: 상대 경로만 허용
            if not nxt.startswith("/"):
                nxt = url_for("index")
            return redirect(nxt)
        error = "아이디 또는 비밀번호가 올바르지 않습니다."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
def index():
    return render_template("scanner.html")


@app.route("/config")
def config_page():
    return render_template("config.html")


@app.route("/console")
def console_page():
    return render_template("console.html", ws_available=_WS_AVAILABLE)


# ─ 스캔 ──────────────────────────────────────────────────────────

@app.route("/api/scan")
def api_scan():
    networks_raw = request.args.get("network", "192.168.0.0/23,192.168.1.0/24").strip()
    mode         = request.args.get("mode", "arp").strip()
    known        = _known_map()

    # 쉼표로 구분된 복수 네트워크 지원; 중복 제거 후 병합하여 단일 스캔
    network_list = [n.strip() for n in networks_raw.split(",") if n.strip()]
    if not network_list:
        network_list = ["192.168.0.0/23"]

    def generate():
        yield sse("ping", msg="스캔 시작")
        yield sse("known_map", data=known)
        for net in network_list:
            yield from scan_generator(net, mode)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ─ 알려진 장비 CRUD ──────────────────────────────────────────────

@app.route("/api/known", methods=["GET"])
def api_known_list():
    rows = list(_known_map().values())
    rows.sort(key=lambda r: r["added_at"], reverse=True)
    return jsonify(rows)


@app.route("/api/known", methods=["POST"])
def api_known_add():
    d  = request.json or {}
    ip = (d.get("ip") or "").strip()
    if not ip:
        return jsonify({"error": "ip required"}), 400
    with sqlite3.connect(DB_PATH) as c:
        c.execute("""
            INSERT INTO known_devices (ip, category, label, note, mac, service_tag, model, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now','localtime'))
            ON CONFLICT(ip) DO UPDATE SET
                category=excluded.category, label=excluded.label,
                note=excluded.note, mac=excluded.mac,
                service_tag=CASE WHEN excluded.service_tag!='' THEN excluded.service_tag ELSE service_tag END,
                model=CASE WHEN excluded.model!='' THEN excluded.model ELSE model END,
                updated_at=excluded.updated_at
        """, (ip, d.get("category", "infra"), d.get("label", ""),
              d.get("note", ""), d.get("mac", ""),
              d.get("service_tag", ""), d.get("model", "")))
    return jsonify({"ok": True})


@app.route("/api/known/<path:ip>", methods=["DELETE"])
def api_known_del(ip):
    with sqlite3.connect(DB_PATH) as c:
        c.execute("DELETE FROM known_devices WHERE ip=?", (ip,))
    return jsonify({"ok": True})


# ─ 자격증명 CRUD ─────────────────────────────────────────────────

@app.route("/api/credentials", methods=["GET"])
def api_creds_list():
    try:
        with sqlite3.connect(DB_PATH) as c:
            rows = c.execute(
                "SELECT ip, username, note, updated_at FROM credentials ORDER BY ip"
            ).fetchall()
        return jsonify([{"ip": r[0], "username": r[1],
                          "note": r[2], "updated_at": r[3]} for r in rows])
    except Exception:
        return jsonify([])


@app.route("/api/credentials", methods=["POST"])
def api_creds_save():
    d  = request.json or {}
    ip = (d.get("ip") or "").strip()
    if not ip:
        return jsonify({"error": "ip required"}), 400
    with sqlite3.connect(DB_PATH) as c:
        c.execute("""
            INSERT INTO credentials (ip, username, password, note, updated_at)
            VALUES (?, ?, ?, ?, datetime('now','localtime'))
            ON CONFLICT(ip) DO UPDATE SET
                username=excluded.username, password=excluded.password,
                note=excluded.note, updated_at=excluded.updated_at
        """, (ip, d.get("username", "root"), d.get("password", ""), d.get("note", "")))
    return jsonify({"ok": True})


@app.route("/api/credentials/<path:ip>", methods=["DELETE"])
def api_creds_del(ip):
    with sqlite3.connect(DB_PATH) as c:
        c.execute("DELETE FROM credentials WHERE ip=?", (ip,))
    return jsonify({"ok": True})


# ─ 프리셋 CRUD ───────────────────────────────────────────────────

@app.route("/api/presets", methods=["GET"])
def api_presets_list():
    try:
        with sqlite3.connect(DB_PATH) as c:
            rows = c.execute("""
                SELECT id, name, description, category, engine,
                       variables, commands, builtin, created_at
                FROM presets ORDER BY builtin DESC, category, name
            """).fetchall()
        return jsonify([{
            "id": r[0], "name": r[1], "description": r[2],
            "category": r[3], "engine": r[4],
            "variables": json.loads(r[5]), "commands": json.loads(r[6]),
            "builtin": bool(r[7]), "created_at": r[8],
        } for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/presets", methods=["POST"])
def api_presets_create():
    d = request.json or {}
    with sqlite3.connect(DB_PATH) as c:
        cur = c.execute("""
            INSERT INTO presets (name, description, category, engine, variables, commands, builtin)
            VALUES (?, ?, ?, ?, ?, ?, 0)
        """, (d.get("name", "새 프리셋"),
              d.get("description", ""),
              d.get("category", "custom"),
              d.get("engine", "both"),
              json.dumps(d.get("variables", [])),
              json.dumps(d.get("commands", []))))
    return jsonify({"ok": True, "id": cur.lastrowid})


@app.route("/api/presets/<int:pid>", methods=["PUT"])
def api_presets_update(pid):
    d = request.json or {}
    with sqlite3.connect(DB_PATH) as c:
        c.execute("""
            UPDATE presets SET name=?, description=?, category=?,
                engine=?, variables=?, commands=?
            WHERE id=?
        """, (d.get("name"), d.get("description", ""),
              d.get("category", "custom"), d.get("engine", "both"),
              json.dumps(d.get("variables", [])),
              json.dumps(d.get("commands", [])), pid))
    return jsonify({"ok": True})


@app.route("/api/presets/<int:pid>", methods=["DELETE"])
def api_presets_delete(pid):
    with sqlite3.connect(DB_PATH) as c:
        row = c.execute("SELECT builtin FROM presets WHERE id=?", (pid,)).fetchone()
        if row and row[0]:
            return jsonify({"error": "내장 프리셋은 삭제할 수 없습니다"}), 400
        c.execute("DELETE FROM presets WHERE id=?", (pid,))
    return jsonify({"ok": True})


# ─ 실행 ──────────────────────────────────────────────────────────

@app.route("/api/execute", methods=["POST"])
def api_execute_start():
    payload = request.json or {}
    job_id  = uuid.uuid4().hex[:10]
    q       = queue.Queue()
    _active_jobs[job_id] = q
    threading.Thread(target=_run_execution, args=(q, payload), daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/api/execute/<job_id>/stream")
def api_execute_stream(job_id):
    q = _active_jobs.get(job_id)
    if not q:
        return jsonify({"error": "job not found"}), 404

    def generate():
        while True:
            msg = q.get()
            if msg is None:
                break
            yield msg
        _active_jobs.pop(job_id, None)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ─ 실행 이력 ─────────────────────────────────────────────────────

@app.route("/api/exec_history", methods=["GET"])
def api_exec_history():
    limit = int(request.args.get("limit", 50))
    try:
        with sqlite3.connect(DB_PATH) as c:
            rows = c.execute("""
                SELECT id, preset_name, engine, targets, variables, results, executed_at
                FROM exec_history ORDER BY executed_at DESC LIMIT ?
            """, (limit,)).fetchall()
        return jsonify([{
            "id": r[0], "preset_name": r[1], "engine": r[2],
            "targets": json.loads(r[3]),
            "variables": json.loads(r[4]),
            "results": json.loads(r[5]),
            "executed_at": r[6],
        } for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─ SCP 프로파일 CRUD ────────────────────────────────────────────

@app.route("/api/scp_profiles", methods=["GET"])
def api_scp_list():
    try:
        with sqlite3.connect(DB_PATH) as c:
            rows = c.execute(
                "SELECT id,name,description,component,created_at FROM scp_profiles ORDER BY created_at DESC"
            ).fetchall()
        return jsonify([{"id":r[0],"name":r[1],"description":r[2],"component":r[3],"created_at":r[4]} for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/scp_profiles", methods=["POST"])
def api_scp_create():
    d = request.json or {}
    name    = (d.get("name") or "").strip()
    content = (d.get("content") or "").strip()
    if not name or not content:
        return jsonify({"error": "name과 content 필수"}), 400
    try:
        with sqlite3.connect(DB_PATH) as c:
            cur = c.execute(
                "INSERT INTO scp_profiles (name,description,content,component) VALUES (?,?,?,?)",
                (name, d.get("description",""), content, d.get("component","ALL"))
            )
        return jsonify({"id": cur.lastrowid})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/scp_profiles/<int:pid>", methods=["GET"])
def api_scp_get(pid):
    try:
        with sqlite3.connect(DB_PATH) as c:
            row = c.execute(
                "SELECT id,name,description,content,component,created_at FROM scp_profiles WHERE id=?", (pid,)
            ).fetchone()
        if not row:
            return jsonify({"error": "없음"}), 404
        return jsonify({"id":row[0],"name":row[1],"description":row[2],"content":row[3],"component":row[4],"created_at":row[5]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/scp_profiles/<int:pid>", methods=["PUT"])
def api_scp_update(pid):
    d = request.json or {}
    try:
        with sqlite3.connect(DB_PATH) as c:
            c.execute(
                "UPDATE scp_profiles SET name=?,description=?,content=?,component=? WHERE id=?",
                (d.get("name",""), d.get("description",""), d.get("content",""), d.get("component","ALL"), pid)
            )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/scp_profiles/<int:pid>", methods=["DELETE"])
def api_scp_delete(pid):
    try:
        with sqlite3.connect(DB_PATH) as c:
            c.execute("DELETE FROM scp_profiles WHERE id=?", (pid,))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─ SCP Export (iDRAC → 앱) ──────────────────────────────────────

def _discover_scp_export_url(ip, username, password):
    """Redfish Managers를 조회해 ExportSystemConfiguration 액션 URL을 반환."""
    try:
        mr = _req.get(f"https://{ip}/redfish/v1/Managers",
                      auth=(username, password), verify=False, timeout=15)
        if mr.status_code == 200:
            members = mr.json().get("Members", [])
            for m in members:
                member_path = m.get("@odata.id", "")
                if not member_path:
                    continue
                dr = _req.get(f"https://{ip}{member_path}",
                              auth=(username, password), verify=False, timeout=15)
                if dr.status_code != 200:
                    continue
                oem_actions = dr.json().get("Actions", {}).get("Oem", {})
                for key, val in oem_actions.items():
                    if "ExportSystemConfiguration" in key:
                        target = val.get("target", "")
                        if target:
                            return target   # e.g. /redfish/v1/Managers/iDRAC.Embedded.1/Actions/Oem/DellManager.ExportSystemConfiguration
    except Exception:
        pass
    # fallback
    return "/redfish/v1/Managers/iDRAC.Embedded.1/Actions/Oem/DellManager.ExportSystemConfiguration"


@app.route("/api/scp_export", methods=["POST"])
def api_scp_export():
    import time
    d         = request.json or {}
    ip        = (d.get("ip") or "").strip()
    username  = d.get("username", "root")
    password  = d.get("password", "calvin")
    component = d.get("component", "ALL")
    if not ip:
        return jsonify({"error": "ip 필수"}), 400
    try:
        action_path = _discover_scp_export_url(ip, username, password)
        url = f"https://{ip}{action_path}"
        # component가 콤마 구분이면 배열로 변환
        comp_val = [c.strip() for c in component.split(",")] if "," in component else component
        r = _req.post(url, auth=(username, password),
            json={"ExportFormat":"XML","ExportUse":"Clone",
                  "ShareParameters":{"Target": comp_val, "ShareType":"Local"}},
            verify=False, timeout=30)
        if r.status_code not in (200, 202):
            return jsonify({"error": f"HTTP {r.status_code}: {r.text[:300]}"}), 400

        loc    = r.headers.get("Location","")
        job_id = loc.split("/")[-1] if loc else ""
        if not job_id:
            return jsonify({"error": "Job ID 없음"}), 500

        for _ in range(72):          # 최대 6분 폴링
            time.sleep(5)
            jr = _req.get(f"https://{ip}/redfish/v1/TaskService/Tasks/{job_id}",
                auth=(username, password), verify=False, timeout=15)
            if jr.status_code != 200:
                continue
            task  = jr.json()
            state = task.get("TaskState", "")
            if state == "Completed":
                for msg in task.get("Messages", []):
                    text = msg.get("Message", "")
                    if "<SystemConfiguration" in text:
                        return jsonify({"xml": text})
                return jsonify({"error": "XML 없음 (작업 완료됐으나 내용 없음)"}), 500
            if state in ("Exception","Killed","Failed"):
                msgs = " | ".join(m.get("Message","") for m in task.get("Messages",[]))
                return jsonify({"error": f"Export 실패: {msgs}"}), 500

        return jsonify({"error": "Export 타임아웃 (6분 초과)"}), 408
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─ SCP 일괄 적용 ─────────────────────────────────────────────────

def _run_scp_apply(q: queue.Queue, payload: dict):
    import time
    profile_id    = payload.get("profile_id")
    targets       = payload.get("targets", [])
    var_values    = payload.get("variables", {})
    component     = payload.get("component", "ALL")
    shutdown_type = payload.get("shutdown_type", "Graceful")
    use_common    = payload.get("use_common_cred", False)
    common_cred   = payload.get("common_cred", {})

    try:
        with sqlite3.connect(DB_PATH) as c:
            row = c.execute("SELECT name,content FROM scp_profiles WHERE id=?", (profile_id,)).fetchone()
        if not row:
            q.put(sse("error", msg="SCP 프로파일을 찾을 수 없습니다"))
            q.put(None)
            return
        profile_name, xml_template = row
    except Exception as e:
        q.put(sse("error", msg=str(e)))
        q.put(None)
        return

    all_results = {}
    xml_content = substitute_vars(xml_template, var_values)

    for ip in targets:
        cred = common_cred if use_common else _get_cred(ip)
        ip_results = []
        q.put(sse("target_start", ip=ip))

        try:
            url = f"https://{ip}/redfish/v1/Managers/iDRAC.Embedded.1/Actions/Oem/DellManager.ImportSystemConfiguration"
            resp = _req.post(url, auth=(cred["username"], cred["password"]),
                json={"ImportBuffer": xml_content, "ShutdownType": shutdown_type,
                      "TimeToWait": 300, "Target": component},
                verify=False, timeout=30)

            if resp.status_code in (200, 202):
                loc    = resp.headers.get("Location","")
                job_id = loc.split("/")[-1] if loc else ""
                step = {"engine":"scp","cmd":"Import 요청","ok":True,"output":f"Job: {job_id}"}
                ip_results.append(step)
                q.put(sse("step_done", ip=ip, step=step))

                if job_id:
                    for _ in range(72):
                        time.sleep(5)
                        try:
                            jr = _req.get(f"https://{ip}/redfish/v1/TaskService/Tasks/{job_id}",
                                auth=(cred["username"], cred["password"]), verify=False, timeout=15)
                            if jr.status_code == 200:
                                task  = jr.json()
                                state = task.get("TaskState","Running")
                                pct   = task.get("PercentComplete", 0)
                                if state not in ("Completed","Exception","Killed","Failed"):
                                    progress = {"engine":"scp","cmd":f"진행 중 ({pct}%)","ok":True,"output":state}
                                    q.put(sse("step_done", ip=ip, step=progress))
                                    continue
                                ok   = (state == "Completed")
                                msgs = "\n".join(m.get("Message","") for m in task.get("Messages",[]))
                                done_step = {"engine":"scp","cmd":"완료","ok":ok,"output":msgs or state}
                                ip_results.append(done_step)
                                q.put(sse("step_done", ip=ip, step=done_step))
                                break
                        except Exception as pe:
                            err_step = {"engine":"scp","cmd":"폴링 오류","ok":False,"output":str(pe)}
                            ip_results.append(err_step)
                            q.put(sse("step_done", ip=ip, step=err_step))
                            break
            else:
                err_msg = resp.text[:400]
                try:
                    err_msg = resp.json().get("error",{}).get("message", err_msg)
                except Exception:
                    pass
                err_step = {"engine":"scp","cmd":"Import 요청","ok":False,"output":f"HTTP {resp.status_code}: {err_msg}"}
                ip_results.append(err_step)
                q.put(sse("step_done", ip=ip, step=err_step))
        except Exception as e:
            err_step = {"engine":"scp","cmd":"오류","ok":False,"output":str(e)}
            ip_results.append(err_step)
            q.put(sse("step_done", ip=ip, step=err_step))

        all_results[ip] = ip_results
        q.put(sse("target_done", ip=ip))

    q.put(sse("done", results=all_results, ts=datetime.now().strftime("%H:%M:%S")))
    q.put(None)


@app.route("/api/scp_apply", methods=["POST"])
def api_scp_apply():
    payload = request.json or {}
    if not payload.get("profile_id") or not payload.get("targets"):
        return jsonify({"error": "profile_id, targets 필수"}), 400
    job_id = str(uuid.uuid4())
    q: queue.Queue = queue.Queue()
    _active_jobs[job_id] = q
    threading.Thread(target=_run_scp_apply, args=(q, payload), daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/api/scp_apply/<job_id>/stream")
def api_scp_apply_stream(job_id):
    q = _active_jobs.get(job_id)
    if not q:
        return Response("data: {\"event\":\"error\",\"msg\":\"job not found\"}\n\n",
                        mimetype="text/event-stream")
    def generate():
        while True:
            msg = q.get()
            if msg is None:
                _active_jobs.pop(job_id, None)
                break
            yield msg
    return Response(stream_with_context(generate()),
                    mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})


# ─ 하드웨어 상세 정보 ────────────────────────────────────────────

@app.route("/api/hw_info/<path:ip>", methods=["GET"])
def api_hw_info(ip):
    refresh = request.args.get("refresh") == "1"
    want_fw = request.args.get("fw") == "1"

    if not refresh:
        try:
            with sqlite3.connect(DB_PATH) as c:
                row = c.execute(
                    "SELECT data, updated_at FROM hw_info WHERE ip=?", (ip,)
                ).fetchone()
            if row:
                data = json.loads(row[0])
                data["cached"] = True
                data["cached_at"] = row[1]
                # 펌웨어 목록을 원하는데 캐시에 없으면 새로 조회
                if not (want_fw and not data.get("firmware")):
                    return jsonify(data)
        except Exception:
            pass

    data = _fetch_hw_detail(ip, want_fw=want_fw)
    if data.get("ok"):
        try:
            with sqlite3.connect(DB_PATH) as c:
                c.execute(
                    "INSERT INTO hw_info (ip, data, updated_at) "
                    "VALUES (?, ?, datetime('now','localtime')) "
                    "ON CONFLICT(ip) DO UPDATE SET data=excluded.data, "
                    "updated_at=excluded.updated_at",
                    (ip, json.dumps(data)),
                )
        except Exception:
            pass
    return jsonify(data)


# ─ 펌웨어 업데이트 ───────────────────────────────────────────────

@app.route("/api/fw_update", methods=["POST"])
def api_fw_update_start():
    try:
        targets = json.loads(request.form.get("targets", "[]"))
    except Exception:
        targets = []
    if not targets:
        return jsonify({"error": "대상 iDRAC를 선택하세요"}), 400

    payload = {
        "targets": targets,
        "apply_time": request.form.get("apply_time", "Immediate"),
        "image_uri": request.form.get("image_uri", "").strip(),
        "transfer_protocol": request.form.get("transfer_protocol", "HTTP"),
        "share_user": request.form.get("share_user", ""),
        "share_pass": request.form.get("share_pass", ""),
        "use_common_cred": request.form.get("use_common_cred") == "1",
        "common_cred": {
            "username": request.form.get("cc_user", "root"),
            "password": request.form.get("cc_pass", ""),
        },
    }

    file_path, file_name, cleanup = "", "", True
    f = request.files.get("file")
    lib_path = request.form.get("lib_path", "").strip()
    if f and f.filename:
        os.makedirs(_FW_TMP_DIR, exist_ok=True)
        file_name = os.path.basename(f.filename)
        file_path = os.path.join(_FW_TMP_DIR, f"{uuid.uuid4().hex}_{file_name}")
        f.save(file_path)
    elif lib_path:
        resolved = _resolve_fw_lib_path(lib_path)
        if not resolved:
            return jsonify({"error": f"라이브러리에서 파일을 찾을 수 없습니다: {lib_path}"}), 400
        file_path = resolved
        file_name = os.path.basename(resolved)
        cleanup = False                      # 공유 원본 파일은 삭제하지 않음
    elif not payload["image_uri"]:
        return jsonify({"error": "펌웨어 파일 · 라이브러리 파일 · 이미지 URI 중 하나가 필요합니다"}), 400

    job_id = uuid.uuid4().hex[:10]
    q: queue.Queue = queue.Queue()
    _active_jobs[job_id] = q
    threading.Thread(target=_run_fw_update,
                     args=(q, payload, file_path, file_name, cleanup), daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/api/fw_files", methods=["GET"])
def api_fw_files():
    root = _fw_lib_root()
    if not root or not os.path.isdir(root):
        return jsonify({"configured": False, "dir": _FW_LIB_DIR, "files": []})

    q = (request.args.get("q") or "").lower()
    files = []
    for cur, dirs, names in os.walk(root):
        dirs[:] = [d for d in dirs if not d.startswith(".")]
        for nm in names:
            if not nm.lower().endswith(_FW_EXTS):
                continue
            full = os.path.join(cur, nm)
            rel = os.path.relpath(full, root).replace("\\", "/")
            if q and q not in rel.lower():
                continue
            try:
                st = os.stat(full)
            except OSError:
                continue
            files.append({
                "name": nm,
                "path": rel,
                "size_mb": round(st.st_size / 1e6, 1),
                "mtime": datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
        if len(files) >= 3000:
            break
    files.sort(key=lambda x: x["mtime"], reverse=True)
    return jsonify({"configured": True, "dir": _FW_LIB_DIR,
                    "count": len(files), "files": files[:1000]})


@app.route("/api/fw_update/<job_id>/stream")
def api_fw_update_stream(job_id):
    q = _active_jobs.get(job_id)
    if not q:
        return jsonify({"error": "job not found"}), 404

    def generate():
        while True:
            msg = q.get()
            if msg is None:
                break
            yield msg
        _active_jobs.pop(job_id, None)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ─ XLSX 내보내기 ─────────────────────────────────────────────────

@app.route("/api/export/xlsx", methods=["POST"])
def api_export_xlsx():
    from io import BytesIO
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return jsonify({"error": "openpyxl 미설치 — 이미지를 재빌드하세요"}), 500

    d = request.json or {}
    hosts = d.get("hosts", [])
    fetch_hw = bool(d.get("fetch_hw"))
    fetch_fw = bool(d.get("fetch_fw"))
    known = _known_map()
    ips = [h.get("ip") for h in hosts if h.get("ip")]

    # ── hw_info 수집: 캐시 우선, 필요 시 실시간 조회 ──
    hw_by_ip: dict = {}
    if ips:
        try:
            with sqlite3.connect(DB_PATH) as c:
                ph = ",".join("?" * len(ips))
                for ip, data_s, upd in c.execute(
                        f"SELECT ip, data, updated_at FROM hw_info WHERE ip IN ({ph})", ips):
                    try:
                        obj = json.loads(data_s)
                        obj["_updated"] = upd
                        hw_by_ip[ip] = obj
                    except Exception:
                        pass
        except Exception:
            pass
    if fetch_hw:
        budget = 40
        for ip in ips:
            if budget <= 0:
                break
            cur = hw_by_ip.get(ip)
            if cur and (not fetch_fw or cur.get("firmware")):
                continue
            budget -= 1
            info = _fetch_hw_detail(ip, want_fw=fetch_fw)
            if info.get("ok"):
                try:
                    with sqlite3.connect(DB_PATH) as c:
                        c.execute(
                            "INSERT INTO hw_info (ip, data, updated_at) "
                            "VALUES (?, ?, datetime('now','localtime')) "
                            "ON CONFLICT(ip) DO UPDATE SET data=excluded.data, "
                            "updated_at=excluded.updated_at",
                            (ip, json.dumps(info)))
                except Exception:
                    pass
                info["_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                hw_by_ip[ip] = info

    def _hw(ip):
        return hw_by_ip.get(ip) or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "스캔결과"
    ws.append(["IP", "서비스태그", "모델", "MAC", "열린 포트", "iDRAC", "Dell",
               "스캔 방법", "등록 분류", "등록 레이블", "메모"])
    for h in hosts:
        kn = known.get(h.get("ip", ""), {})
        ws.append([
            h.get("ip", ""),
            h.get("service_tag", ""),
            h.get("model", ""),
            h.get("mac", ""),
            " ".join(str(p) for p in h.get("open_ports", [])),
            "Y" if h.get("is_idrac") else "",
            "Y" if h.get("is_dell") else "",
            h.get("method", ""),
            kn.get("category", ""),
            kn.get("label", ""),
            h.get("note", "") or kn.get("note", ""),
        ])

    def _sheet(title, headers):
        sh = wb.create_sheet(title)
        sh.append(headers)
        return sh

    # ── HW요약 (호스트 1행) ──
    sm = _sheet("HW요약", ["IP", "레이블", "모델", "서비스태그", "BIOS", "CPU",
                           "메모리(GiB)", "디스크 수", "PSU 수", "iDRAC FW", "상태", "조회시각"])
    for h in hosts:
        ip = h.get("ip", "")
        kn = known.get(ip, {})
        hw = _hw(ip)
        if not hw:
            sm.append([ip, kn.get("label", ""), h.get("model", ""),
                       h.get("service_tag", ""), "", "", "", "", "", "", "(HW 미조회)", ""])
            continue
        s = hw.get("system", {}) or {}
        idr = hw.get("idrac", {}) or {}
        pw = hw.get("power", {}) or {}
        sm.append([
            ip, kn.get("label", ""), s.get("model", ""), s.get("service_tag", ""),
            s.get("bios_version", ""),
            (f'{s.get("cpu_count", "")} x {s.get("cpu_model", "")}').strip(" x"),
            s.get("mem_total_gib", ""),
            sum(len(x.get("drives", [])) for x in hw.get("storage", [])),
            len(pw.get("supplies", [])),
            idr.get("firmware_version", ""),
            s.get("health", ""), hw.get("_updated", ""),
        ])

    # ── CPU ──
    sc = _sheet("CPU", ["IP", "소켓", "모델", "코어", "스레드", "속도(MHz)", "상태"])
    for h in hosts:
        ip = h.get("ip", "")
        for p in _hw(ip).get("processors", []):
            sc.append([ip, p.get("id", ""), p.get("model", ""), p.get("cores", ""),
                       p.get("threads", ""), p.get("speed_mhz", ""),
                       p.get("health") or p.get("state", "")])

    # ── 메모리 ──
    smem = _sheet("메모리", ["IP", "슬롯", "용량(GiB)", "속도", "타입", "제조사",
                             "파트번호", "시리얼", "상태"])
    for h in hosts:
        ip = h.get("ip", "")
        for m in _hw(ip).get("memory", []):
            smem.append([ip, m.get("locator", ""), m.get("capacity_gib", ""),
                         m.get("speed_mhz", ""), m.get("type", ""), m.get("manufacturer", ""),
                         m.get("part_number", ""), m.get("serial", ""),
                         m.get("health") or m.get("state", "")])

    # ── 디스크 ──
    sd = _sheet("디스크", ["IP", "컨트롤러", "이름", "용량(GB)", "미디어", "프로토콜",
                           "모델", "시리얼", "상태"])
    for h in hosts:
        ip = h.get("ip", "")
        for ctrl in _hw(ip).get("storage", []):
            cname = ctrl.get("model") or ctrl.get("name", "")
            for x in ctrl.get("drives", []):
                sd.append([ip, cname, x.get("name", ""), x.get("capacity_gb", ""),
                           x.get("media", ""), x.get("proto", ""), x.get("model", ""),
                           x.get("serial", ""), x.get("health") or x.get("state", "")])

    # ── 네트워크 ──
    sn = _sheet("네트워크", ["IP", "ID", "이름", "MAC", "속도(Mbps)", "링크", "상태"])
    for h in hosts:
        ip = h.get("ip", "")
        for n in _hw(ip).get("network", []):
            sn.append([ip, n.get("id", ""), n.get("name", ""), n.get("mac", ""),
                       n.get("speed_mbps", ""), n.get("link", ""), n.get("health", "")])

    # ── 전원(PSU) ──
    sp = _sheet("전원", ["IP", "이름", "모델", "시리얼", "펌웨어", "용량(W)", "입력(W)", "상태"])
    for h in hosts:
        ip = h.get("ip", "")
        for p in (_hw(ip).get("power", {}) or {}).get("supplies", []):
            sp.append([ip, p.get("name", ""), p.get("model", ""), p.get("serial", ""),
                       p.get("firmware", ""), p.get("capacity_w", ""), p.get("input_w", ""),
                       p.get("health") or p.get("state", "")])

    # ── 펌웨어 인벤토리 (데이터 있을 때만) ──
    if any(_hw(h.get("ip", "")).get("firmware") for h in hosts):
        sf = _sheet("펌웨어", ["IP", "구성요소", "버전", "업데이트 가능"])
        for h in hosts:
            ip = h.get("ip", "")
            for f in _hw(ip).get("firmware", []):
                up = f.get("updateable")
                sf.append([ip, f.get("name", ""), f.get("version", ""),
                           "예" if up is True else "아니오" if up is False else ""])

    # 헤더만 남은 상세 시트 제거
    for nm in ["CPU", "메모리", "디스크", "네트워크", "전원"]:
        if nm in wb.sheetnames and wb[nm].max_row <= 1:
            del wb[nm]

    for sh in wb.worksheets:
        for cell in sh[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="374151")
        sh.freeze_panes = "A2"
        for col in sh.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            sh.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 48)

    bio = BytesIO()
    wb.save(bio)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="idrac_scan_{ts}.xlsx"'},
    )


# ─ 시스템 정보 ───────────────────────────────────────────────────

@app.route("/api/system_info", methods=["GET"])
def api_system_info():
    return jsonify({
        "racadm_available": racadm_available(),
        "racadm_path": shutil.which("racadm") or "",
        "ws_available": _WS_AVAILABLE,
        "fw_lib_dir": _FW_LIB_DIR,
        "fw_lib_ready": bool(_fw_lib_root() and os.path.isdir(_fw_lib_root())),
    })


# ─ 웹 SSH 콘솔 (페이지 이동 시에도 세션 유지) ───────────────────

_SSH_SESSIONS: dict = {}
_SSH_LOCK = threading.Lock()
_SSH_BUF_MAX     = 256 * 1024      # 재접속 시 되살릴 스크롤백 최대 크기
_SSH_IDLE_MAX    = 30 * 60        # ws 분리 상태로 이 시간 지나면 세션 종료
_SSH_DEAD_GRACE  = 90            # 원격 셸 종료 후 버퍼 보존 시간
_SSH_MAX         = 20


class _SSHSession:
    def __init__(self, sid, cli, chan, meta):
        self.sid = sid
        self.cli = cli
        self.chan = chan
        self.meta = meta
        self.buf = bytearray()
        self.lock = threading.Lock()
        self.ws = None
        self.stop = threading.Event()
        self.dead_at = 0.0
        self.detached_at = time.time()

    def _send_ws(self, obj):
        w = self.ws
        if w is None:
            return
        try:
            w.send(json.dumps(obj))
        except Exception:
            self.ws = None
            self.detached_at = time.time()

    def append(self, data: bytes):
        with self.lock:
            self.buf.extend(data)
            if len(self.buf) > _SSH_BUF_MAX:
                del self.buf[:len(self.buf) - _SSH_BUF_MAX]
            self._send_ws({"type": "data", "data": data.decode("utf-8", "replace")})

    def close(self):
        self.stop.set()
        try:
            self.chan.close()
        except Exception:
            pass
        try:
            self.cli.close()
        except Exception:
            pass


def _ssh_pump(sess: "_SSHSession"):
    ch = sess.chan
    while not sess.stop.is_set():
        try:
            data = ch.recv(8192)
            if not data:
                break
            sess.append(data)
        except socket.timeout:
            time.sleep(0.02)
        except Exception:
            break
    sess.dead_at = time.time()
    with sess.lock:
        sess._send_ws({"type": "status", "data": "closed"})


def _ssh_gc():
    while True:
        time.sleep(30)
        now = time.time()
        with _SSH_LOCK:
            for sid, s in list(_SSH_SESSIONS.items()):
                dead = s.dead_at and now - s.dead_at > _SSH_DEAD_GRACE
                idle = s.ws is None and s.detached_at and now - s.detached_at > _SSH_IDLE_MAX
                if dead or idle:
                    s.close()
                    _SSH_SESSIONS.pop(sid, None)


def _register_ws_routes():
    if not sock:
        return

    threading.Thread(target=_ssh_gc, daemon=True).start()

    @sock.route("/ws/ssh")
    def ws_ssh(ws):                     # noqa: ANN001
        import paramiko

        raw = ws.receive(timeout=30)
        if not raw:
            return
        try:
            cfg = json.loads(raw)
        except Exception:
            ws.send(json.dumps({"type": "error", "data": "잘못된 접속 정보"}))
            return

        sid = cfg.get("sid") or ""
        sess = None
        if sid:
            with _SSH_LOCK:
                sess = _SSH_SESSIONS.get(sid)
            if sess is None:
                ws.send(json.dumps({"type": "expired",
                                    "data": "세션이 만료되었습니다. 다시 접속하세요."}))
                return

        # ── 신규 세션 ──
        if sess is None:
            ip   = (cfg.get("ip") or "").strip()
            user = cfg.get("username") or "root"
            pw   = cfg.get("password") or ""
            port = int(cfg.get("port") or 22)
            cols = int(cfg.get("cols") or 120)
            rows = int(cfg.get("rows") or 32)
            if not ip:
                ws.send(json.dumps({"type": "error", "data": "IP 주소가 필요합니다"}))
                return
            with _SSH_LOCK:
                if len(_SSH_SESSIONS) >= _SSH_MAX:
                    ws.send(json.dumps({"type": "error",
                                        "data": "동시 SSH 세션 한도를 초과했습니다"}))
                    return
            cli = paramiko.SSHClient()
            cli.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            try:
                cli.connect(ip, port=port, username=user, password=pw,
                            timeout=12, banner_timeout=15, auth_timeout=15,
                            look_for_keys=False, allow_agent=False)
                chan = cli.invoke_shell(term="xterm", width=cols, height=rows)
            except Exception as e:
                ws.send(json.dumps({"type": "error", "data": f"SSH 접속 실패: {e}"}))
                try:
                    cli.close()
                except Exception:
                    pass
                return
            chan.settimeout(0.0)
            sid = uuid.uuid4().hex
            sess = _SSHSession(sid, cli, chan, {"ip": ip, "user": user, "port": port})
            with _SSH_LOCK:
                _SSH_SESSIONS[sid] = sess
            threading.Thread(target=_ssh_pump, args=(sess,), daemon=True).start()
            ws.send(json.dumps({"type": "session", "sid": sid,
                                "ip": ip, "user": user}))

        # ── ws 연결(재)부착 ──
        prev = sess.ws
        if prev is not None and prev is not ws:
            try:
                prev.send(json.dumps({"type": "status", "data": "detached"}))
            except Exception:
                pass
        with sess.lock:
            try:
                ws.send(json.dumps({"type": "attached",
                                    "ip": sess.meta.get("ip", ""),
                                    "user": sess.meta.get("user", "")}))
                if sess.buf:
                    ws.send(json.dumps({"type": "data",
                                        "data": bytes(sess.buf).decode("utf-8", "replace")}))
            except Exception:
                return
            sess.ws = ws
            sess.detached_at = 0.0
        if sess.stop.is_set() or sess.dead_at:
            try:
                ws.send(json.dumps({"type": "status", "data": "closed"}))
            except Exception:
                pass

        # ── 입력 루프 ──
        try:
            while not sess.stop.is_set():
                msg = ws.receive(timeout=1)
                if msg is None:
                    if sess.chan.exit_status_ready():
                        break
                    continue
                try:
                    m = json.loads(msg)
                except Exception:
                    continue
                t = m.get("type")
                if t == "data":
                    try:
                        sess.chan.send(m.get("data", ""))
                    except Exception:
                        break
                elif t == "resize":
                    try:
                        sess.chan.resize_pty(width=int(m.get("cols", 120)),
                                             height=int(m.get("rows", 32)))
                    except Exception:
                        pass
                elif t == "close":
                    sess.close()
                    with _SSH_LOCK:
                        _SSH_SESSIONS.pop(sess.sid, None)
                    try:
                        ws.send(json.dumps({"type": "status", "data": "closed"}))
                    except Exception:
                        pass
                    break
        except Exception:
            pass
        finally:
            # 페이지 이동 등으로 ws만 끊긴 경우 → 세션·셸은 살려둔다
            if sess.ws is ws:
                sess.ws = None
                sess.detached_at = time.time()


_register_ws_routes()


# ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5001))

    # HTTPS 설정: SSL_CERTFILE + SSL_KEYFILE 환경변수 또는 SSL_SELF_SIGNED=1
    ssl_ctx = None
    certfile = os.environ.get("SSL_CERTFILE", "")
    keyfile  = os.environ.get("SSL_KEYFILE",  "")
    if certfile and keyfile and os.path.exists(certfile) and os.path.exists(keyfile):
        ssl_ctx = (certfile, keyfile)
        print(f"[HTTPS] 인증서 로드: {certfile}")
    elif os.environ.get("SSL_SELF_SIGNED", ""):
        cert_path = "/tmp/idrac_cert.pem"
        key_path  = "/tmp/idrac_key.pem"
        if not (os.path.exists(cert_path) and os.path.exists(key_path)):
            try:
                subprocess.run([
                    "openssl", "req", "-x509", "-newkey", "rsa:2048",
                    "-keyout", key_path, "-out", cert_path,
                    "-days", "3650", "-nodes",
                    "-subj", "/CN=idrac-manager"
                ], check=True, capture_output=True)
                print("[HTTPS] 자체 서명 인증서 생성 완료")
            except Exception as e:
                print(f"[HTTPS] 인증서 생성 실패: {e}")
        if os.path.exists(cert_path) and os.path.exists(key_path):
            ssl_ctx = (cert_path, key_path)

    app.run(host="0.0.0.0", port=port, debug=False, threaded=True,
            ssl_context=ssl_ctx)
